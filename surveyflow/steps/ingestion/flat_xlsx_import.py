"""Convert a flat "Question" + "Data" 2-sheet Excel workbook into the
standard MCP pair (``definition.json`` + ``data_export.csv``), so the rest
of the ingestion pipeline (``metadata_parser.parse_metadata`` +
``export_parser.parse_export_csv``/``convert_export_to_rawdata``) runs
completely unchanged.

Why this exists
----------------
Some surveys arrive as a single ``.xlsx`` with 2 sheets instead of the
standard ``get_survey_definition`` JSON + QMe "code"-format
``data_export.csv``:

- **"Question"** sheet: one row per question OR per sub-item (choice /
  matrix row / rank-pool item / MA choice). Columns: ``Name of items``,
  ``Question type`` (``SA``/``MA``/``FT``/``RANKING``), ``Question(Matrix)``
  (shared question text when the row is one of several sub-items),
  ``Question(Normal)`` (the question text, or the row/choice label when
  ``Question(Matrix)`` is set), then choice texts in the numbered columns.
- **"Data"** sheet: DOES match the standard QMe "code" export layout (the
  header row's first cell is ``"Approve"``) - just embedded as a sheet
  instead of a standalone CSV.

Grouping heuristics (how a "Question" sheet gets turned into definition
questions)
----------------------------------------------------------------------
For every column in the **real** Data-sheet header (the source of truth for
what columns actually exist - some Question-sheet sub-item groups collapse
into fewer real columns, see below):

1. Strip a trailing ``_{digits}`` or ``_o{digits}`` suffix to get a "base"
   label and group columns sharing that base.
2. A group of 1 column, found by its exact name in the Question sheet
   → standalone SA/FT question.
3. A group of 1 column, NOT found by its exact name, but where the
   Question sheet lists MULTIPLE rows whose name starts with
   ``"{base}_"`` → the real export collapsed a multi-item group into one
   column but QMe's own multiplenumber sub-label patch
   (``export_parser._patch_multiplenumber_headers``) will re-expand it into
   per-category ``{base}_n{k}`` columns at parse time - modeled as
   ``multiplenumber``.
4. A group of 2+ columns, all type ``RANKING`` in the Question sheet → a
   ranking question; every row in the group repeats the SAME full item pool
   (verified against several example surveys), so choices are read from
   just the first row.
5. A group of 2+ columns, all type ``MA`` → a multiple-choice question;
   ``Question(Normal)`` is the SHARED question text (repeated per row) and
   the actual choice label is the row's first choice-column value.
6. A group of 2+ columns, all type ``SA``, sharing the same
   ``Question(Matrix)`` text → a ``Matrix_SA`` question: rows keyed by the
   numeric suffix, row label = ``Question(Normal)``, columns = the choice
   list (verified identical across the group's rows).
7. ``_o{n}`` suffixed columns mark an "other-specify" free-text field tied
   to choice code ``n`` of the group's flat choices - flagged
   ``is_other: true`` in ``choices_i18n``.

Known limitations (surfaced via the returned ``warnings`` list - always
review them):

- Only the 4 answer types actually observed so far (``SA``, ``MA``, ``FT``,
  ``RANKING``) are handled. Any other value in the "Question type" column
  is skipped with a warning rather than guessed at.
- No vi/en translation is available in the source sheet - question_i18n /
  choices_i18n duplicate the same text into both "vi" and "en".
- Interviewer-only fields (question text matching ``[interviewer`` /
  ``do not ask``) and photo-upload fields (label ending in ``_PHOTO``) are
  excluded from ``definition["questions"]`` entirely (never asked to a
  respondent / not meaningful in a data table) - this is a naming-pattern
  heuristic, not a guarantee; check the warnings if a survey's photo/
  interviewer fields use a different naming convention.
"""
from __future__ import annotations

import csv
import json
import re
from pathlib import Path

_SUFFIX_OTHER = re.compile(r"^(.+)_o(\d+)$")
_SUFFIX_IDX   = re.compile(r"^(.+)_(\d+)$")

_HANDLED_QTYPES = {"SA", "MA", "FT", "RANKING"}

# Standard QMe "code"-format report columns - appear on every export
# regardless of survey content, never treated as questions.
QME_SYSTEM_COLS: set[str] = {
    "Approve", "Reject", "Re-do request", "Reason to reject", "Memo", "Edited",
    "Edited by", "Edited ratio", "No.", "Date", "ID", "Country", "Channel", "Method",
    "Login ID", "User name", "IP address (Public user)", "Store ID", "Store Code",
    "Store name", "District", "Ward", "Store address", "Group 1", "Group 2", "Group 3",
    "Area group", "Region 2", "Manager", "Telephone number", "Contact person", "Email",
    "Others 1", "Others 2", "Others 3", "Others 4", "Check in", "Store Latitude",
    "Store Longitude", "User Latitude", "User Longitude", "Check out", "Distance",
    "Task duration", "Speeder Flag", "Speed Ratio", "Speed Percentile (%)",
    "Total Duration (s)", "Num Questions Seen", "Estimated sec", "QC Score Total",
    "SL Points", "OE Points", "Speeder Points", "Median Points", "Comments",
}

_INTERVIEWER_ONLY_RE = re.compile(r"\[interviewer|do not ask", re.IGNORECASE)


def _is_photo_label(label: str) -> bool:
    return label.upper().endswith("_PHOTO") or label.upper() == "PHOTO"


def _base_and_kind(col: str) -> tuple[str, "int | None", str]:
    m = _SUFFIX_OTHER.match(col)
    if m:
        return m.group(1), int(m.group(2)), "other"
    m = _SUFFIX_IDX.match(col)
    if m:
        return m.group(1), int(m.group(2)), "idx"
    return col, None, "plain"


def read_question_sheet(ws) -> dict[str, dict]:
    """openpyxl worksheet (the "Question" sheet) -> {name: item dict}."""
    items: dict[str, dict] = {}
    rows = list(ws.iter_rows(values_only=True))
    for r in rows[1:]:
        if not r or not r[0]:
            continue
        name = str(r[0]).strip()
        items[name] = {
            "name": name,
            "qtype": (str(r[1]).strip() if len(r) > 1 and r[1] else ""),
            "matrix_text": (str(r[2]).strip() if len(r) > 2 and r[2] else ""),
            "normal_text": (str(r[3]).strip() if len(r) > 3 and r[3] else ""),
            "choices": [str(c).strip() for c in r[4:] if c not in (None, "")],
        }
    return items


def dump_data_sheet_to_csv(ws, out_path: Path) -> None:
    """Write the "Data" sheet to CSV, sanitizing embedded newlines (Alt+Enter
    cells) that would otherwise break the line-based header/data-row
    detection in ``export_parser.parse_export_csv`` (which splits on
    physical newlines, not real CSV row boundaries)."""

    def _clean(v: object) -> str:
        if v is None:
            return ""
        return str(v).replace("\r\n", " ").replace("\n", " ").replace("\r", " ")

    with open(out_path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            w.writerow([_clean(v) for v in row])


def _find_real_header(csv_path: Path) -> list[str]:
    """Return the column list from the row whose first field is 'Approve'."""
    with open(csv_path, encoding="utf-8-sig") as f:
        for line in f:
            row = next(csv.reader([line]), [])
            if row and row[0].strip() == "Approve":
                return [c.strip() for c in row]
    raise RuntimeError(f"{csv_path}: header row (first field == 'Approve') not found")


def build_definition_from_flat_questions(
    q_items: dict[str, dict],
    real_header: list[str],
    survey_id: int = 1,
    survey_title: str = "",
) -> tuple[dict, list[str]]:
    """Core heuristic converter — see module docstring for the grouping rules.

    Returns (definition_dict, warnings) where definition_dict has the same
    shape as ``get_survey_definition``'s response
    (``{"survey": {...}, "questions": [...]}``), consumable directly by
    ``metadata_parser.parse_metadata``.
    """
    groups: dict[str, list[tuple[str, "int|None", str]]] = {}
    order: list[str] = []
    for col in real_header:
        if not col or col in QME_SYSTEM_COLS:
            continue
        base, idx, kind = _base_and_kind(col)
        groups.setdefault(base, []).append((col, idx, kind))
        if base not in order:
            order.append(base)

    questions: list[dict] = []
    warnings: list[str] = []
    qid = 100000
    pos = 0

    for base in order:
        cols = groups[base]
        other_cols = [(c, idx) for c, idx, kind in cols if kind == "other"]
        main_cols  = [(c, idx, kind) for c, idx, kind in cols if kind != "other"]
        if not main_cols:
            warnings.append(f"[skip] {base}: only 'other' column(s), no main sibling")
            continue

        # ── Single column (no siblings under this base) ──────────────────
        # Applies regardless of whether the column name happens to end in a
        # numeric suffix (e.g. "Q53_3") - a trailing digit alone doesn't
        # mean matrix/MA grouping if there's no sibling column to group with.
        if len(main_cols) == 1:
            col = main_cols[0][0]
            item = q_items.get(col)

            if item is not None and _is_photo_label(col):
                warnings.append(f"[exclude photo] {col}")
                continue
            if item is not None and _INTERVIEWER_ONLY_RE.search(item["normal_text"]):
                warnings.append(f"[exclude interviewer-only] {col}")
                continue

            if item is None:
                # Collapsed group (real export shows ONE column, but the
                # Question sheet lists several "{col}_{n}" sub-items) —
                # export_parser's own multiplenumber sub-label patch will
                # re-expand it into per-category "{col}_n{k}" columns at
                # parse time. Modeled as multiplenumber.
                cand_items = [(k, v) for k, v in q_items.items()
                              if k == col or k.startswith(col + "_")]
                if not cand_items:
                    warnings.append(f"[skip] {col}: no match in Question sheet")
                    continue

                def _suffix_num(k: str) -> int:
                    m = re.search(r"_(\d+)$", k)
                    return int(m.group(1)) if m else 0

                cand_items.sort(key=lambda kv: _suffix_num(kv[0]))
                text = cand_items[0][1]["matrix_text"] or cand_items[0][1]["normal_text"]
                choices_i18n = {
                    str(_suffix_num(k) or (i + 1)): {"vi": v["normal_text"], "en": v["normal_text"]}
                    for i, (k, v) in enumerate(cand_items)
                }
                warnings.append(
                    f"[assume multiplenumber] {col}: collapsed to 1 real column but "
                    f"Question sheet lists {len(cand_items)} sub-items - verify once "
                    "real data confirms export_parser expands it"
                )
                pos += 1
                qid += 1
                questions.append({
                    "position": pos, "question_id": qid, "label": col,
                    "type": 1, "input_type": 100,
                    "question_i18n": {"vi": text, "en": text},
                    "mandatory": False, "status": 1, "choices_i18n": choices_i18n,
                })
                continue

            qtype = item["qtype"]
            text  = item["normal_text"]
            if qtype not in _HANDLED_QTYPES:
                warnings.append(f"[skip] {col}: unhandled qtype={qtype!r}")
                continue

            pos += 1
            qid += 1
            if qtype == "FT":
                q = {
                    "position": pos, "question_id": qid, "label": col,
                    "type": 1, "input_type": 0,
                    "question_i18n": {"vi": text, "en": text},
                    "mandatory": False, "status": 1, "choices_i18n": {},
                }
            elif qtype == "SA":
                choices_i18n = {
                    str(i + 1): {"vi": c, "en": c}
                    for i, c in enumerate(item["choices"])
                }
                q = {
                    "position": pos, "question_id": qid, "label": col,
                    "type": 2, "input_type": 0,
                    "question_i18n": {"vi": text, "en": text},
                    "mandatory": False, "status": 1, "choices_i18n": choices_i18n,
                }
            else:
                warnings.append(f"[skip] {col}: standalone qtype={qtype} unexpected "
                                 "for a single column (MA/RANKING normally need siblings)")
                pos -= 1
                qid -= 1
                continue

            _apply_other(q, other_cols, warnings)
            questions.append(q)
            continue

        # ── Grouped columns (MA / Matrix_SA / ranking) ──────────────────
        items_found = [(c, idx, q_items.get(c)) for c, idx, kind in main_cols]
        missing = [c for c, idx, it in items_found if it is None]
        if missing:
            warnings.append(f"[skip] {base}: grouped columns missing from "
                             f"Question sheet: {missing}")
            continue

        qtypes = {it["qtype"] for _, _, it in items_found}
        pos += 1
        qid += 1

        if qtypes == {"RANKING"}:
            first_item = items_found[0][2]
            text = first_item["normal_text"]
            choices_i18n = {
                str(i + 1): {"vi": c, "en": c}
                for i, c in enumerate(first_item["choices"])
            }
            questions.append({
                "position": pos, "question_id": qid, "label": base,
                "type": 6, "input_type": 0,
                "question_i18n": {"vi": text, "en": text},
                "mandatory": False, "status": 1, "choices_i18n": choices_i18n,
            })
            continue

        if qtypes == {"MA"}:
            # normal_text (col3) is the SHARED question text, repeated on
            # every sub-item row; the actual choice label is choices[0].
            text = items_found[0][2]["normal_text"]
            items_sorted = sorted(items_found, key=lambda t: t[1])
            choices_i18n = {
                str(idx): {
                    "vi": (it["choices"][0] if it["choices"] else it["normal_text"]),
                    "en": (it["choices"][0] if it["choices"] else it["normal_text"]),
                }
                for _, idx, it in items_sorted
            }
            q = {
                "position": pos, "question_id": qid, "label": base,
                "type": 3, "input_type": 0,
                "question_i18n": {"vi": text, "en": text},
                "mandatory": False, "status": 1, "choices_i18n": choices_i18n,
            }
            _apply_other(q, other_cols, warnings)
            questions.append(q)
            continue

        if qtypes == {"SA"}:
            # Matrix_SA: matrix_text shared question, normal_text = row
            # label, choices shared across all rows.
            first_item = items_found[0][2]
            matrix_text = first_item["matrix_text"] or first_item["normal_text"]
            rows_map = {str(idx): it["normal_text"] for _, idx, it in items_found}
            columns_i18n = {
                str(i + 1): {"vi": c, "en": c}
                for i, c in enumerate(first_item["choices"])
            }
            questions.append({
                "position": pos, "question_id": qid, "label": base,
                "type": 4, "input_type": 0,
                "question_i18n": {"vi": matrix_text, "en": matrix_text},
                "mandatory": False, "status": 1,
                "choices_i18n": {"rows": rows_map, "columns": columns_i18n},
            })
            continue

        pos -= 1
        qid -= 1
        warnings.append(f"[skip] {base}: mixed/unhandled qtypes={qtypes}")

    definition = {
        "survey": {
            "survey_id": survey_id,
            "title": survey_title,
            "english_title": survey_title,
            "status": "active",
            "start_date": "",
            "end_date": "",
        },
        "questions": questions,
    }
    return definition, warnings


def _apply_other(q: dict, other_cols: list[tuple[str, int]], warnings: list[str]) -> None:
    ci = q.get("choices_i18n", {})
    for other_col, code in other_cols:
        key = str(code)
        if key in ci:
            ci[key]["is_other"] = True
        else:
            warnings.append(
                f"[warn] {other_col}: is_other code {code} not found in "
                f"{q.get('label')}'s choices_i18n"
            )


def convert_xlsx_to_mcp(
    xlsx_path: "str | Path",
    mcp_dir: "str | Path",
    survey_title: str = "",
    survey_id: int = 1,
) -> dict:
    """Read *xlsx_path* (a "Question"+"Data" 2-sheet workbook) and write
    ``definition.json`` + ``data_export.csv`` into *mcp_dir*, ready for the
    standard ``surveyflow-run --mcp-dir ... --export-csv ...`` ingestion
    flow (or the ``--xlsx-input`` shortcut, which calls this directly).

    Returns ``{"definition_path", "data_export_path", "warnings": [...]}``.
    Always review ``warnings`` — they flag every skipped/assumed column.
    """
    import openpyxl

    xlsx_path = Path(xlsx_path)
    mcp_dir   = Path(mcp_dir)
    mcp_dir.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    missing_sheets = {"Question", "Data"} - set(wb.sheetnames)
    if missing_sheets:
        raise ValueError(
            f"{xlsx_path}: expected sheets 'Question' and 'Data', "
            f"missing {sorted(missing_sheets)} (found {wb.sheetnames})"
        )

    q_items = read_question_sheet(wb["Question"])

    data_export_path = mcp_dir / "data_export.csv"
    dump_data_sheet_to_csv(wb["Data"], data_export_path)

    real_header = _find_real_header(data_export_path)
    definition, warnings = build_definition_from_flat_questions(
        q_items, real_header,
        survey_id=survey_id,
        survey_title=survey_title or xlsx_path.stem,
    )

    definition_path = mcp_dir / "definition.json"
    definition_path.write_text(
        json.dumps(definition, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    return {
        "definition_path":  str(definition_path),
        "data_export_path": str(data_export_path),
        "warnings":          warnings,
    }
