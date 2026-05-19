"""TableStep: datatable.json + rawdata + metadata → datatable.xlsx (multi-sheet)."""
from __future__ import annotations

import logging
import time
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from surveyflow.core.base import Step
from surveyflow.steps.table.banner_builder import BannerColumn, build_banner, nest_banner_with_matrix_rows
from surveyflow.steps.table.table_generator import StubBlock, StubRow, RowGroupBlock, RankingBlock, compute_table

logger = logging.getLogger(__name__)

# ── Fills ──────────────────────────────────────────────────────────────────────
_F_GROUP   = PatternFill("solid", fgColor="1F4E79")   # navy  — group row
_F_MID     = PatternFill("solid", fgColor="2E75B6")   # blue  — mid / sub_mid rows
_F_DETAIL  = PatternFill("solid", fgColor="BDD7EE")   # light blue — detail row
_F_STAT    = PatternFill("solid", fgColor="EBF3FB")   # very light blue — stat rows
_F_SIG_HDR   = PatternFill("solid", fgColor="E2EFDA")   # light green — sig col header
_F_GROUP_ROW = PatternFill("solid", fgColor="BDD7EE")   # same as detail — stub group summary

# ── Fonts ──────────────────────────────────────────────────────────────────────
_WHITE_BOLD = Font(color="FFFFFF", bold=True, size=10)
_BOLD       = Font(bold=True, size=10)
_DARK_BOLD  = Font(color="1F4E79", bold=True, size=10)
_RED        = Font(color="FF0000", size=10)
_NORMAL     = Font(size=10)
_GREY       = Font(color="595959", size=9)

# ── Alignments ─────────────────────────────────────────────────────────────────
_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
_L = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_R = Alignment(horizontal="right",  vertical="center")

# ── Borders (black) ────────────────────────────────────────────────────────────
_S_THIN  = Side(style="thin",   color="000000")
_S_MED   = Side(style="medium", color="000000")
_S_THICK = Side(style="thick",  color="000000")

def _brd(thick_left: bool = False, thick_bottom: bool = False) -> Border:
    """Build a cell border with optional thick left (group sep) and thick bottom (section sep)."""
    return Border(
        left   = _S_THICK if thick_left   else _S_THIN,
        right  = _S_THIN,
        top    = _S_THIN,
        bottom = _S_MED   if thick_bottom else _S_THIN,
    )

_BORDER = _brd()   # plain thin border (backwards-compat alias)

_STAT_TYPES = {"t2b", "b2b", "mean", "std", "se"}

# ── Cell helpers ───────────────────────────────────────────────────────────────

def _set(ws, row, col, val=None, *, font=None, fill=None, align=None,
         border=None, num_fmt=None):
    cell = ws.cell(row=row, column=col)
    if val is not None:
        cell.value = val
    if font:    cell.font          = font
    if fill:    cell.fill          = fill
    if align:   cell.alignment     = align
    if border:  cell.border        = border
    if num_fmt: cell.number_format = num_fmt


def _merge(ws, r1, c1, r2, c2):
    if r1 == r2 and c1 == c2:
        return
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)


# ── Banner column layout ───────────────────────────────────────────────────────

def _banner_layout(banner_cols: list[BannerColumn], show_sig: bool) -> list[dict]:
    """
    Return a flat list of column-slot dicts, one per Excel data column.

    Fields per slot
    ---------------
    kind           : "data" | "sig"
    banner         : BannerColumn
    bc_index       : int   — position in banner_cols, used to index StubRow dicts
    first_in_group : bool  — True on the first column of each new banner group
                             → render a thick left border to visually separate groups
    """
    slots: list[dict] = []
    prev_group: str | None = None
    for bc_idx, bc in enumerate(banner_cols):
        first = bc.group_label != prev_group
        prev_group = bc.group_label
        slots.append({
            "kind": "data", "banner": bc,
            "bc_index": bc_idx, "first_in_group": first,
        })
        if show_sig:
            slots.append({
                "kind": "sig", "banner": bc,
                "bc_index": bc_idx, "first_in_group": False,
            })
    return slots


# ── Sheet writer ───────────────────────────────────────────────────────────────

def _write_sheet(
    ws,
    config: dict,
    table_cfg: dict,
    banner_cols: list[BannerColumn],
    blocks: list[StubBlock | RowGroupBlock],
) -> None:
    cell_content = table_cfg.get("cell_content", "percentage")   # "count" | "percentage"
    show_sig     = table_cfg.get("show_sig", False)
    sig_cfg      = config.get("significance_test", {"enabled": False})

    slots   = _banner_layout(banner_cols, show_sig)
    n_slots = len(slots)
    DATA_COL = 5   # col E is first data column

    # ── Column widths ────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 6    # code col
    ws.column_dimensions["D"].width = 30   # answer label col
    for i, slot in enumerate(slots):
        ltr = get_column_letter(DATA_COL + i)
        ws.column_dimensions[ltr].width = 6 if slot["kind"] == "sig" else 13

    # ── Row 1: title ─────────────────────────────────────────────────
    _merge(ws, 1, 1, 1, DATA_COL + n_slots - 1)
    _set(ws, 1, 1, config.get("title", "Data Table"),
         font=Font(bold=True, size=13, color="1F4E79"), align=_L)
    ws.row_dimensions[1].height = 22

    # ── Row 2: sub_title (or spacer) ─────────────────────────────────
    sub_title = config.get("sub_title", "")
    if sub_title:
        _merge(ws, 2, 1, 2, DATA_COL + n_slots - 1)
        _set(ws, 2, 1, sub_title,
             font=Font(bold=True, size=11, color="2E75B6"), align=_L)
        ws.row_dimensions[2].height = 18
    else:
        ws.row_dimensions[2].height = 5

    # ── Row 3: cell content info ─────────────────────────────────────
    _set(ws, 3, 1, f"Cell content: {cell_content}", font=_NORMAL)
    ws.row_dimensions[3].height = 16

    # ── Row 4: sig test info (method + levels) ───────────────────────
    if show_sig and sig_cfg.get("enabled"):
        lvls   = sig_cfg.get("levels", [95])
        method = sig_cfg.get("method", "independent").capitalize()
        parts  = []
        if 95 in lvls: parts.append("Uppercase = 95%")
        if 90 in lvls: parts.append("lowercase = 90%")
        _set(ws, 4, 1,
             f"Sig test: {', '.join(parts)}    |    Method: {method}",
             font=_NORMAL)
    ws.row_dimensions[4].height = 16

    # ── Row 5: banner GROUP labels ───────────────────────────────────
    group_spans: list[tuple[str, int, int, bool]] = []   # label, c1, c2, first
    for i, slot in enumerate(slots):
        col   = DATA_COL + i
        label = slot["banner"].group_label
        if group_spans and group_spans[-1][0] == label:
            group_spans[-1] = (label, group_spans[-1][1], col, group_spans[-1][3])
        else:
            first = len(group_spans) == 0   # no thick for first group's left edge
            group_spans.append((label, col, col, first))

    for idx, (label, c1, c2, _) in enumerate(group_spans):
        thick = idx > 0   # thick left border between groups (not on the very first)
        _merge(ws, 5, c1, 5, c2)
        _set(ws, 5, c1, label,
             font=_WHITE_BOLD, fill=_F_GROUP, align=_C, border=_brd(thick))
    ws.row_dimensions[5].height = 20

    # ── Dynamic row numbering ─────────────────────────────────────────────────
    # max_levels = max number of intermediate level rows across all banner columns.
    # 0 levels → plain:    group(5) / detail(6) / letter(7)
    # N levels → N-level:  group(5) / level0(6) / … / level(N-1)(5+N) / detail(6+N) / letter(7+N)
    max_levels = max((len(bc.level_labels) for bc in banner_cols), default=0)
    DETAIL_ROW = 6 + max_levels
    LETTER_ROW = DETAIL_ROW + 1
    DATA_START = (LETTER_ROW + 1) if show_sig else (DETAIL_ROW + 1)

    # ── Level rows (one per intermediate level) ────────────────────────────────
    # At level row i, columns that have a label at that depth show it merged across
    # all consecutive sibling slots (same group + same level prefix up to i).
    # Columns with fewer levels show an empty filled cell.
    # Fill: outermost level = navy (_F_GROUP), all inner levels = blue (_F_MID).
    for level_idx in range(max_levels):
        level_row = 6 + level_idx
        fill      = _F_GROUP if level_idx == 0 else _F_MID
        i = 0
        while i < len(slots):
            slot  = slots[i]
            bc    = slot["banner"]
            c1    = DATA_COL + i
            thick = slot["first_in_group"] and i > 0

            # Find end of merge run: same group + same level prefix up to level_idx
            j = i + 1
            while j < len(slots):
                other = slots[j]["banner"]
                if (other.group_label != bc.group_label or
                        other.level_labels[:level_idx + 1] != bc.level_labels[:level_idx + 1]):
                    break
                j += 1
            c2 = DATA_COL + j - 1

            if len(bc.level_labels) > level_idx:
                _merge(ws, level_row, c1, level_row, c2)
                _set(ws, level_row, c1, bc.level_labels[level_idx],
                     font=_WHITE_BOLD, fill=fill, align=_C, border=_brd(thick))
            else:
                _merge(ws, level_row, c1, level_row, c2)
                _set(ws, level_row, c1, fill=fill, border=_brd(thick))
            i = j
        ws.row_dimensions[level_row].height = 18

    # ── Detail row: ALL columns show subgroup_label ───────────────────────
    detail_thick_bot = not show_sig
    i = 0
    while i < len(slots):
        slot  = slots[i]
        bc    = slot["banner"]
        c1    = DATA_COL + i
        j     = i
        while j < len(slots) and slots[j]["banner"] is bc:
            j += 1
        thick = slot["first_in_group"] and i > 0
        _merge(ws, DETAIL_ROW, c1, DETAIL_ROW, DATA_COL + j - 1)
        _set(ws, DETAIL_ROW, c1, bc.subgroup_label,
             font=_DARK_BOLD, fill=_F_DETAIL, align=_C,
             border=_brd(thick, thick_bottom=detail_thick_bot))
        i = j
    ws.row_dimensions[DETAIL_ROW].height = 18

    # ── Letter row ────────────────────────────────────────────────────────
    if show_sig:
        for i, slot in enumerate(slots):
            col   = DATA_COL + i
            thick = slot["first_in_group"] and i > 0
            if slot["kind"] == "data":
                _set(ws, LETTER_ROW, col, slot["banner"].letter,
                     font=_BOLD, fill=_F_DETAIL, align=_C,
                     border=_brd(thick, thick_bottom=True))
            else:
                _set(ws, LETTER_ROW, col, "Sig",
                     font=_GREY, fill=_F_STAT, align=_C,
                     border=_brd(False, thick_bottom=True))
        ws.row_dimensions[LETTER_ROW].height = 16
    else:
        ws.row_dimensions[LETTER_ROW].height = 0

    # ── Stub rows ────────────────────────────────────────────────────
    cur = DATA_START

    def _write_block(block: StubBlock, cur: int, is_sub: bool = False) -> int:
        """Write one StubBlock. Returns updated cur row number."""
        if not block.rows:
            return cur

        base_row  = next((r for r in block.rows if r.row_type == "base"), None)
        data_rows = [r for r in block.rows if r.row_type != "base"]

        # Question header row (doubles as Base row)
        hdr_fill = _F_STAT if is_sub else _F_MID
        hdr_font = _BOLD if is_sub else _WHITE_BOLD
        _set(ws, cur, 1, block.question_code,
             font=hdr_font, fill=hdr_fill, align=_C, border=_brd(thick_bottom=True))
        _set(ws, cur, 2, f"{block.question_label} ({block.answer_type})",
             font=hdr_font, fill=hdr_fill, align=_L, border=_brd(thick_bottom=True))
        _set(ws, cur, 3, "Code",
             font=hdr_font, fill=hdr_fill, align=_C, border=_brd(thick_bottom=True))
        _set(ws, cur, 4, "Base" if base_row else "",
             font=hdr_font, fill=hdr_fill, align=_L, border=_brd(thick_bottom=True))

        for i, slot in enumerate(slots):
            col   = DATA_COL + i
            thick = slot["first_in_group"] and i > 0
            if slot["kind"] == "sig" or base_row is None:
                _set(ws, cur, col, fill=hdr_fill, border=_brd(thick, thick_bottom=True))
                continue
            bc_idx = slot["bc_index"]
            val    = int(base_row.counts.get(bc_idx, 0))
            _set(ws, cur, col, val,
                 font=hdr_font, fill=hdr_fill,
                 align=_R, border=_brd(thick, thick_bottom=True), num_fmt="0")

        ws.row_dimensions[cur].height = 17
        cur += 1

        # Data rows
        n_data = len(data_rows)
        for row_idx, stub_row in enumerate(data_rows):
            r           = cur
            is_stat     = stub_row.row_type in _STAT_TYPES
            is_group    = stub_row.row_type == "group"
            is_last_row = row_idx == n_data - 1

            _row_fill = (
                _F_STAT      if is_stat  else
                _F_GROUP_ROW if is_group else
                None
            )
            _lbl_font = (
                _DARK_BOLD if is_stat  else
                _BOLD      if is_group else
                _NORMAL
            )

            _set(ws, r, 1, border=_brd(thick_bottom=is_last_row))
            _set(ws, r, 2, border=_brd(thick_bottom=is_last_row))

            _code_val = None
            if not is_stat and not is_group and stub_row.code is not None:
                try:
                    _code_val = int(stub_row.code)
                except (ValueError, TypeError):
                    _code_val = stub_row.code
            _set(ws, r, 3, _code_val,
                 font=_NORMAL, fill=_row_fill,
                 align=_C, border=_brd(thick_bottom=is_last_row))

            _set(ws, r, 4, stub_row.label,
                 font=_lbl_font, fill=_row_fill,
                 align=_L, border=_brd(thick_bottom=is_last_row))

            for i, slot in enumerate(slots):
                col    = DATA_COL + i
                bc_idx = slot["bc_index"]
                thick  = slot["first_in_group"] and i > 0

                if slot["kind"] == "sig":
                    if stub_row.row_type == "percent" and show_sig:
                        mark = stub_row.sig_marks.get(bc_idx, "")
                        _set(ws, r, col, mark or None,
                             font=_RED if mark else _NORMAL,
                             align=_C, border=_brd(False, thick_bottom=is_last_row))
                    else:
                        _set(ws, r, col, border=_brd(False, thick_bottom=is_last_row))
                    continue

                _computed_stat = stub_row.row_type in ("mean", "std", "se")
                raw = (
                    stub_row.values.get(bc_idx)
                    if (cell_content != "count" or _computed_stat)
                    else stub_row.counts.get(bc_idx)
                )

                if cell_content == "count" and not _computed_stat:
                    val = int(raw) if raw is not None else 0
                    fmt = "0"
                    fnt = _DARK_BOLD if is_stat else (_BOLD if is_group else _NORMAL)
                elif stub_row.row_type in ("percent", "group"):
                    val = float(raw) if raw is not None else 0.0
                    fmt = "0%"
                    fnt = _BOLD if is_group else _NORMAL
                else:
                    val = float(raw) if raw is not None else 0.0
                    fmt = "0%" if stub_row.row_type in ("t2b", "b2b") else "0.00"
                    fnt = _DARK_BOLD

                _set(ws, r, col, val,
                     font=fnt, fill=_row_fill, align=_R,
                     border=_brd(thick, thick_bottom=is_last_row),
                     num_fmt=fmt)

            ws.row_dimensions[r].height = 15
            cur += 1

        return cur

    for block in blocks:

        if isinstance(block, RowGroupBlock):
            # ── Section header row (navy, full width) ────────────────
            last_col = DATA_COL + len(slots) - 1
            _merge(ws, cur, 1, cur, last_col)
            _set(ws, cur, 1, block.row_label,
                 font=_WHITE_BOLD, fill=_F_GROUP, align=_L,
                 border=_brd(thick_bottom=True))
            ws.row_dimensions[cur].height = 18
            cur += 1
            # Sub-blocks (one per question item) — rendered with subdued header
            for sub_block in block.sub_blocks:
                cur = _write_block(sub_block, cur, is_sub=True)

        elif isinstance(block, RankingBlock):
            if block.mode == "any_rank":
                # Flat MA-style — render the single sub_block as a normal block
                cur = _write_block(block.sub_blocks[0], cur, is_sub=False)
            else:
                # rank_dist — navy section header then one sub-block per rank position
                last_col = DATA_COL + len(slots) - 1
                _merge(ws, cur, 1, cur, last_col)
                _set(ws, cur, 1,
                     f"{block.question_code}  {block.question_label} (ranking)",
                     font=_WHITE_BOLD, fill=_F_GROUP, align=_L,
                     border=_brd(thick_bottom=True))
                ws.row_dimensions[cur].height = 18
                cur += 1
                for sub_block in block.sub_blocks:
                    cur = _write_block(sub_block, cur, is_sub=True)

        else:
            cur = _write_block(block, cur, is_sub=False)

    # freeze panes
    ws.freeze_panes = ws.cell(row=DATA_START, column=DATA_COL)


# ── Helpers ────────────────────────────────────────────────────────────────────

def _build_lookup_maps(metadata: dict) -> tuple[dict[str, str], dict[str, dict]]:
    """Build col_map and q_pos_to_meta from metadata.

    col_map maps question reference → actual df column name:
      - SA / ranking / old semicolon-MA : rawdata_columns[0] (the real df column)
      - Binary-MA (rawdata_columns len > 1): label kept as-is; callers read
        q_meta["rawdata_columns"] directly and route per-choice to the right column.
    """
    col_map: dict[str, str] = {}
    q_pos_to_meta: dict[str, dict] = {}
    for meta in metadata.get("questions", {}).values():
        pos = meta.get("position")
        if pos is None:
            continue
        key          = f"q{pos}"
        label        = meta.get("label") or key
        atype        = meta.get("answer_type", "")
        rawdata_cols = meta.get("rawdata_columns", [])

        # Binary-MA: multiple rawdata_columns (one per choice code).
        # Keep col_map pointing to label; banner / generator handle each column
        # individually via q_meta["rawdata_columns"].
        is_ma_binary = atype == "MA" and len(rawdata_cols) > 1
        primary = label if is_ma_binary else (rawdata_cols[0] if rawdata_cols else label)

        col_map[key]       = primary
        q_pos_to_meta[key] = meta
        if label and label != key:
            col_map[label]       = primary
            q_pos_to_meta[label] = meta
    return col_map, q_pos_to_meta


# ── Step ───────────────────────────────────────────────────────────────────────

class TableStep(Step):
    """
    Context inputs
    --------------
    df               : pd.DataFrame
    metadata         : dict
    datatable_config : list[dict] | dict
    output_dir       : str

    Context outputs (compute)
    -------------------------
    table_results    : list[dict]   JSON-serialisable cross-tab results
    _table_computed  : list[dict]   raw objects for render_xlsx (internal use)

    Context outputs (render_xlsx)
    -----------------------------
    datatable_path   : str
    """

    def compute(self, context: dict[str, Any]) -> dict[str, Any]:
        """Compute cross-tabs for selected configs. No file I/O — returns structured data.

        Filtering
        ---------
        ``context["table_indices"]`` (list[int] | None)
            Indices into the datatable_config array to compute.
            None (default) → compute all configs.
        """
        t_compute_start = time.perf_counter()

        df         = context["df"]
        metadata   = context["metadata"]
        raw_config = context["datatable_config"]

        all_configs: list[dict] = raw_config if isinstance(raw_config, list) else [raw_config]

        # Filter to requested indices (preserve original index for reference)
        indices = context.get("table_indices")   # list[int] | None
        if indices is not None:
            selected = [(i, all_configs[i]) for i in indices if 0 <= i < len(all_configs)]
        else:
            selected = list(enumerate(all_configs))

        col_map, q_pos_to_meta = _build_lookup_maps(metadata)

        computed: list[dict] = []   # raw objects → consumed by render_xlsx
        results:  list[dict] = []   # serialised  → consumed by API / preview

        for cfg_idx, (orig_idx, cfg) in enumerate(selected, 1):
            sub_title  = cfg.get("sub_title", "")
            sig_config = cfg.get("significance_test", {"enabled": False})
            tag = f"[{cfg_idx}/{len(selected)}] {sub_title or 'config'}"

            t0 = time.perf_counter()
            logger.info("%s — building banner …", tag)
            banner_cols = build_banner(cfg, df, col_map=col_map, q_pos_to_meta=q_pos_to_meta)

            bm = cfg.get("banner_matrix")
            if bm:
                banner_cols = nest_banner_with_matrix_rows(
                    base_columns    = banner_cols,
                    matrix_question = bm["question"],
                    df              = df,
                    q_pos_to_meta   = q_pos_to_meta,
                    col_map         = col_map,
                    groups          = bm.get("groups"),
                )
            logger.info("%s — banner: %d columns  (%.2fs)", tag, len(banner_cols),
                        time.perf_counter() - t0)

            t0 = time.perf_counter()
            logger.info("%s — computing cross-tabs …", tag)
            blocks = compute_table(
                stub_configs   = cfg["stub"],
                banner_cols    = banner_cols,
                df             = df,
                metadata       = metadata,
                sig_config     = sig_config,
                col_map        = col_map,
                q_pos_to_meta  = q_pos_to_meta,
            )
            logger.info("%s — cross-tabs: %d blocks  (%.2fs)", tag, len(blocks),
                        time.perf_counter() - t0)

            computed.append({
                "orig_idx":   orig_idx,
                "cfg":        cfg,
                "banner_cols": banner_cols,
                "blocks":     blocks,
            })

        context["_table_computed"] = computed

        # Always build JSON-serialisable table_results so callers (APDP preview,
        # API, or xlsx export) can read structured data without a second compute().
        t0 = time.perf_counter()
        logger.info("Serialising to dict …")
        context["table_results"] = [
            {
                "table_index":       item["orig_idx"],
                "title":             item["cfg"].get("title", ""),
                "sub_title":         item["cfg"].get("sub_title", ""),
                "significance_test": item["cfg"].get("significance_test", {"enabled": False}),
                "banner_cols": [bc.to_dict() for bc in item["banner_cols"]],
                "blocks":      [b.to_dict()  for b  in item["blocks"]],
                "tables":      item["cfg"].get("tables", []),
            }
            for item in computed
        ]
        logger.info("Serialised  (%.2fs)", time.perf_counter() - t0)

        logger.info("compute() done  (%.2fs total)", time.perf_counter() - t_compute_start)
        return context

    def render_xlsx(self, context: dict[str, Any]) -> dict[str, Any]:
        """Write datatable.xlsx.

        Source priority
        ---------------
        1. ``_table_computed`` — Python objects produced in the same process by
           ``compute()``.  Fast: no reconstruction needed.
        2. ``table_results`` — JSON produced by a prior ``compute()`` call,
           e.g. loaded from storage in a separate APDP export job.
           Reconstructed via ``_from_table_results()`` before rendering.
        """
        t_render_start = time.perf_counter()

        if "_table_computed" in context:
            computed = context["_table_computed"]
        elif "table_results" in context:
            logger.info("render_xlsx: reconstructing from table_results JSON …")
            computed = self._from_table_results(context["table_results"])
        else:
            raise ValueError(
                "render_xlsx() requires either '_table_computed' or 'table_results' in context."
            )

        out_dir  = Path(context.get("output_dir", "."))
        out_dir.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        wb.remove(wb.active)

        for item in computed:
            cfg         = item["cfg"]
            banner_cols = item["banner_cols"]
            blocks      = item["blocks"]
            sub_title   = cfg.get("sub_title", "")

            tables = cfg.get("tables", [
                {"sheet": "Table", "cell_content": "percentage", "show_sig": False, "enabled": True}
            ])
            for tbl in tables:
                if not tbl.get("enabled", True):
                    continue
                base_name  = tbl.get("sheet", "Sheet")
                sheet_name = f"{sub_title} - {base_name}" if sub_title else base_name
                t0 = time.perf_counter()
                logger.info("Writing sheet: %s …", sheet_name)
                ws = wb.create_sheet(title=sheet_name)
                _write_sheet(ws, cfg, tbl, banner_cols, blocks)
                logger.info("  → %s done  (%.2fs)", sheet_name, time.perf_counter() - t0)

        output_path = str(out_dir / "datatable.xlsx")
        t0 = time.perf_counter()
        logger.info("Saving workbook …")
        wb.save(output_path)
        logger.info("Saved → %s  (%.2fs)", output_path, time.perf_counter() - t0)
        logger.info("render_xlsx() done  (%.2fs total)", time.perf_counter() - t_render_start)

        context["datatable_path"] = output_path
        return context

    # ── JSON → Python object reconstruction ───────────────────────────────────

    def _from_table_results(self, table_results: list[dict]) -> list[dict]:
        """Reconstruct internal computed list from table_results JSON.

        Called by render_xlsx() when _table_computed is absent — i.e. when the
        APDP export job reads a previously stored table_results.json instead of
        rerunning compute().

        Returns a list in the same shape as _table_computed so that render_xlsx()
        can use it without modification.
        """
        computed: list[dict] = []
        for item in table_results:
            banner_cols = [self._bc_from_dict(d)    for d in item.get("banner_cols", [])]
            blocks      = [self._block_from_dict(d) for d in item.get("blocks", [])]
            cfg = {
                "title":             item.get("title", ""),
                "sub_title":         item.get("sub_title", ""),
                "tables":            item.get("tables", []),
                "significance_test": item.get("significance_test", {"enabled": False}),
            }
            computed.append({
                "orig_idx":    item.get("table_index", 0),
                "cfg":         cfg,
                "banner_cols": banner_cols,
                "blocks":      blocks,
            })
        return computed

    @staticmethod
    def _bc_from_dict(d: dict) -> BannerColumn:
        """Reconstruct BannerColumn from to_dict() output.

        ``mask`` is set to an empty Series — it is only needed during compute(),
        not during rendering, so this is safe for render_xlsx() use.
        """
        return BannerColumn(
            group_label    = d.get("group_label", ""),
            subgroup_label = d.get("label", ""),
            letter         = d.get("letter", ""),
            mask           = pd.Series(dtype=bool),   # dummy — not used in render
            is_total       = d.get("is_total", False),
            level_labels   = d.get("level_labels", []),
        )

    @staticmethod
    def _row_from_dict(d: dict) -> StubRow:
        """Reconstruct StubRow from to_dict() array format → internal dict format."""
        values_arr = d.get("values", [])
        counts_arr = d.get("counts", [])
        sig_arr    = d.get("sig_marks", [])
        return StubRow(
            label     = d.get("label", ""),
            row_type  = d.get("row_type", "percent"),
            code      = d.get("code"),
            values    = {i: float(v) for i, v in enumerate(values_arr)},
            counts    = {i: int(v)   for i, v in enumerate(counts_arr)},
            sig_marks = {i: s for i, s in enumerate(sig_arr) if s},
        )

    @staticmethod
    def _stub_from_dict(d: dict) -> StubBlock:
        """Reconstruct StubBlock from to_dict() output."""
        return StubBlock(
            question_code  = d.get("question", ""),
            question_label = d.get("label", ""),
            answer_type    = d.get("answer_type", ""),
            rows           = [TableStep._row_from_dict(r) for r in d.get("rows", [])],
        )

    @staticmethod
    def _block_from_dict(d: dict) -> StubBlock | RowGroupBlock | RankingBlock:
        """Reconstruct the correct block type from to_dict() output.

        Dispatches on ``d["type"]``:
          "stub"      → StubBlock
          "row_group" → RowGroupBlock  (sub_blocks → StubBlock list)
          "ranking"   → RankingBlock   (sub_blocks → StubBlock list)
        """
        btype = d.get("type", "stub")
        if btype == "row_group":
            return RowGroupBlock(
                row_label  = d.get("row_label", ""),
                row_code   = d.get("row_code", ""),
                sub_blocks = [TableStep._stub_from_dict(b) for b in d.get("sub_blocks", [])],
            )
        if btype == "ranking":
            return RankingBlock(
                question_code  = d.get("question", ""),
                question_label = d.get("label", ""),
                answer_type    = d.get("answer_type", "ranking"),
                mode           = d.get("mode", "rank_dist"),
                sub_blocks     = [TableStep._stub_from_dict(b) for b in d.get("sub_blocks", [])],
            )
        # default → stub
        return TableStep._stub_from_dict(d)

    def run(self, context: dict[str, Any]) -> dict[str, Any]:
        """Backward-compatible: compute + render_xlsx in one call."""
        context = self.compute(context)
        context = self.render_xlsx(context)
        return context
